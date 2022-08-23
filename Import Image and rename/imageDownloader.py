import urllib.request
import csv
import time
import threading
import os
import openpyxl

file = open('Applicant_Rolls_8_16_22.csv')
csvreader = csv.reader(file)
header = []
header = next(csvreader)
isjpg = True
isjpeg , ispng = False, False
isJPG, isJPEG, isPNG = False, False, False

error_applicant_images, error_signature_images = set(), set()

def download_applicant_image(applicant_code, applicant_id, image_format):

    image_url = "http://pwd.teletalk.com.bd/images/"+applicant_code+"/applicant_image"+image_format

    save_image = "applicant_image/"+applicant_code+image_format
    save_image_id = "applicant_image/applicant_image_id/"+applicant_id+image_format
    try:
        urllib.request.urlretrieve(image_url, save_image)
        urllib.request.urlretrieve(image_url, save_image_id)
        print("Done Image")
    except:
        error_applicant_images.add((applicant_code,applicant_id))
        print("Image not Done")
        print("error for image", error_applicant_images)
        #download_applicant_image(applicant_code, applicant_id, "JPG")
        #download_applicant_image(applicant_code, applicant_id, "PNG")



def download_applicant_signature(applicant_code, applicant_id, image_format):

    image_url = "http://pwd.teletalk.com.bd/images/"+applicant_code+"/applicant_signature"+image_format
    save_image = "applicant_signature/"+applicant_code+image_format
    save_image_id = "applicant_signature/applicant_signature_id/"+applicant_id+image_format

    try:
        urllib.request.urlretrieve(image_url, save_image)
        urllib.request.urlretrieve(image_url, save_image_id)
        print("Done Signature")
    except:
        error_signature_images.add((applicant_code,applicant_id))
        print("Signature not Done")
        print("error for signature", error_signature_images)
        #download_applicant_signature(applicant_code, applicant_id, "JPG")



def download_both(applicant_code,applicant_id, image_format):
    if os.path.isfile("applicant_image/"+applicant_code+image_format) and os.path.isfile("applicant_image/applicant_image_id/"+applicant_id+image_format) and os.path.isfile("applicant_signature/"+applicant_code+image_format) and os.path.isfile("applicant_signature/applicant_signature_id/"+applicant_id+image_format):
        print(applicant_code, "and", applicant_id, "already exists.")
    else:
        #threading
        t1 = threading.Thread(download_applicant_image(applicant_code, applicant_id, image_format))
        t2 = threading.Thread(download_applicant_signature(applicant_code, applicant_id, image_format))
        # download_applicant_image(applicant_code, applicant_id, image_format)
        # download_applicant_signature(applicant_code, applicant_id, image_format)
        t1.start()
        t2.start()
        t1.join()
        t2.join()
        print("Done Downloading for", i, applicant_code, applicant_id)


tic = time.process_time()
rows = []
for row in csvreader:
    rows.append(row)
print(len(rows))
for i in range(0, len(rows)):
    applicant_code = rows[i][1]
    applicant_id = rows[i][3]
    print("Start Downloading for", i, applicant_code)
    download_both(applicant_code, applicant_id, ".jpg")

toc = time.process_time()
error_applicant_images_list = list(error_applicant_images)
print(error_applicant_images_list)
print("total applicant numbers:", len(error_applicant_images_list))
error_signature_images_list = list(error_signature_images)
print(error_signature_images_list)
print("total signature number:", len(error_signature_images_list))
error_images = error_applicant_images.union(error_signature_images)
error_images_list = list(error_images)
print(error_images_list)
print("total number:", len(error_images_list))
print("Time Elapsed:", toc-tic)

#Saving errors in worksheet

wb = openpyxl.load_workbook("Different_file_format_images.xlsx")
ws = wb.active

# for i in range(0, len(error_applicant_images_list)):
#     ws.cell(i + 2, 1).value = error_applicant_images_list[i]
#
# for i in range(0, len(error_signature_images_list)):
#     ws.cell(i + 2, 2).value = error_signature_images_list[i]
#jpg Error check
for i in range(0, len(error_images_list)):
    ws.cell(i+2, 5).value = error_images_list[i][0]
    ws.cell(i + 2, 6).value = error_images_list[i][1]
    ws.cell(i + 2, 7).value = f'http://pwd.teletalk.com.bd/images/{error_images_list[i][0]}/'
    if error_images_list[i] in error_applicant_images_list and error_images_list[i] in error_signature_images_list:
        ws.cell(i+2, 1).value = error_images_list[i][0]
        ws.cell(i + 2, 2).value = error_images_list[i][1]
        ws.cell(i + 2, 3).value = error_images_list[i][0]
        ws.cell(i + 2, 4).value = error_images_list[i][1]
        print("Done1")
    elif error_images_list[i] in error_applicant_images_list:
        ws.cell(i + 2, 1).value = error_images_list[i][0]
        ws.cell(i + 2, 2).value = error_images_list[i][1]
        print("Done2")
    elif error_images_list[i] in error_signature_images_list:
        ws.cell(i + 2, 3).value = error_images_list[i][0]
        ws.cell(i + 2, 4).value = error_images_list[i][1]
        print("Done3")
    else:
        print("Value not added")
############
#JPG error check
# for i in range(0, len(error_images_list)):
#     ws.cell(i+2, 10).value = error_images_list[i][0]
#     ws.cell(i + 2, 11).value = error_images_list[i][1]
#     ws.cell(i + 2, 12).value = f'http://pwd.teletalk.com.bd/images/{error_images_list[i][0]}/'
#     if error_images_list[i] in error_applicant_images_list and error_images_list[i] in error_signature_images_list:
#         ws.cell(i+2, 8).value = error_images_list[i][0]
#         ws.cell(i + 2, 9).value = error_images_list[i][0]
#     elif error_images_list[i] in error_applicant_images_list:
#         ws.cell(i + 2, 8).value = error_images_list[i][0]
#     elif error_images_list[i] in error_signature_images_list:
#         ws.cell(i + 2, 9).value = error_images_list[i][0]
############

wb.save("Different_file_format_images.xlsx")
